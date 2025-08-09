"""
FastAPI service for Relay: PDF účetní závěrka (VZZ) → vyplněná Excel šablona
------------------------------------------------------------------------------
Zjednodušený, robustní postup dle domluvy (v2 – lepší párování):
- Do Relay nahráváš už jen relevantní stránky (VZZ) a název PDF obsahuje rok, např. "2023.pdf".
- Běžné účetní období = rok z názvu souboru (R), Minulé účetní období = R-1.
- Šablona: list "List1", sloupec D = názvy položek, řádek 3 sloupce E–J = roky (čísly 2019..2024 aj.).
- Zápis probíhá pouze do roků, které v šabloně skutečně existují.
- V2 zlepšení: čištění štítků (odstranění římských/abecedních prefixů, hvězdiček, uvozovek apod.),
  robustnější fuzzy shoda (WRatio), synonymní slovník, ignorování šumových řádků.
- Přidává LOG sheet s nespárovanými/ignorovanými položkami a informacemi o zpracování.

Endpoint: POST /process  (multipart form)
  - file: PDF s VZZ (název obsahuje rok, např. 2023.pdf)
  - template: XLSX šablona (viz výše)
  - threshold: (volitelné) int 0–100; default 76 (doporučeno 72–82 dle kvality PDF)
  - return_log: (volitelné) bool; default True
Vrací: vyplněný XLSX (Content-Disposition: attachment)
"""

import io, re, unicodedata
from typing import List, Dict, Optional, Tuple
from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import StreamingResponse
import pdfplumber
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import process, fuzz

app = FastAPI(title="Relay PDF→Excel VZZ (CZ)")

YEAR_RX = re.compile(r"(20\d{2})")
LABEL_CURRENT = re.compile(r"\b(b[eě]žn[eé]|aktu[aá]ln[íy]|current)\b", re.I)
LABEL_PRIOR   = re.compile(r"\b(minul[eé]|srovn[aá]vac[ií]|p[řr]edchoz[ií]|prior|comparative)\b", re.I)

VZZ_MARKERS = [
    "VÝKAZ ZISKU A ZTRÁTY", "VÝKAZ ZISKŮ A ZTRÁT", "Výkaz zisku a ztráty", "výkaz zisku a ztráty",
    "výsledovka", "Výsledovka"
]

# --------------------------- Helpers ---------------------------

def nz(s: Optional[str]) -> str:
    return "" if s is None else str(s)

def norm(s: str) -> str:
    s = s.strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r"\s+", " ", s)
    return s.lower()

# Odstranění číslovacích prefixů (I., II., a.1, 1.2.3), odrážek, uvozovek, hvězdiček apod.
NUM_PREFIX_RX = re.compile(r"^(?:[IVXLC]+\.|(?:\d+\.)+|\d+[a-z]?\)|[a-z]\.[0-9]+|[a-z]\)|[a-z]\.)\s*", re.I)
JUNK_RX = re.compile(r"[\"'`´•·\*]+")
LETTERS_RX = re.compile(r"[a-zA-Z]\w{2,}")

def clean_label(s: str) -> str:
    s = nz(s)
    s = JUNK_RX.sub(" ", s)
    s = NUM_PREFIX_RX.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

# Synonyma – rozšiřuj postupně dle svých PDF
SYN = {
    "trzby z prodeje vlastnich vyrobku a sluzeb": "tržby z prodeje vlastních výrobků a služeb",
    "trzby za prodej zbozi": "tržby za prodej zboží",
    "zmena stavu zasob": "změna stavu zásob vlastní činnosti",
    "aktivace": "aktivace",
    "naklady vynalozene na prodane zbozi": "náklady vynaložené na prodané zboží",
    "osobni naklady": "osobní náklady",
    "odpisy": "odpisy",
    "dan z prijmu": "daň z příjmů",
}

def apply_synonyms(s: str) -> str:
    return SYN.get(norm(s), s)

# Číslo s podporou účetních závorek a různých oddělovačů

def to_number(cell: str) -> Optional[float]:
    if cell is None:
        return None
    s = str(cell).strip()
    if s == "":
        return None
    s = s.replace("\xa0"," ").replace("\u202f"," ")
    s = s.replace(" ", "")
    s = s.replace("−","-")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9\.\-\(\)]", "", s)
    if s in ("", "-", "--"):
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except:
        return None

# ------------------------ Template reading ---------------------

def collect_template_structure(wb_bytes: bytes):
    wb = load_workbook(io.BytesIO(wb_bytes))
    ws: Worksheet = wb.active  # očekáváme "List1"
    header_row = 3
    year_cols: Dict[int, str] = {}
    years = []
    for col_letter in ["E","F","G","H","I","J"]:
        val = ws[f"{col_letter}{header_row}"].value
        y = None
        if isinstance(val, (int, float)) and 2000 <= int(val) <= 2100:
            y = int(val)
        else:
            m = YEAR_RX.search(str(val) if val is not None else "")
            if m:
                y = int(m.group(1))
        if y:
            year_cols[y] = col_letter
            years.append(y)
    labels: Dict[str, int] = {}
    for r in range(4, ws.max_row+1):
        v = ws[f"D{r}"] .value
        if v is None or str(v).strip()=="":
            continue
        labels[norm(clean_label(str(v)))] = r
    return wb, ws, year_cols, sorted(years), labels

# ------------------------ PDF utilities ------------------------

def detect_vzz_pages(pdf: pdfplumber.PDF) -> List[int]:
    hits = []
    for i, page in enumerate(pdf.pages):
        txt = page.extract_text() or ""
        if any(m.lower() in txt.lower() for m in VZZ_MARKERS):
            hits.append(i)
    if not hits:
        # už posíláš vyříznuté stránky – ber vše
        return list(range(len(pdf.pages)))
    start = hits[0]
    end = min(len(pdf.pages)-1, start+4)
    return list(range(start, end+1))

def extract_tables_from_pages(pdf: pdfplumber.PDF, pages: List[int]) -> List[List[List[str]]]:
    all_tables = []
    for i in pages:
        page = pdf.pages[i]
        tables = page.extract_tables() or []
        cleaned = []
        for t in tables:
            rows = []
            for row in t or []:
                if not row:
                    continue
                rows.append([nz(c) for c in row])
            if rows:
                cleaned.append(rows)
        if not cleaned:
            text = page.extract_text() or ""
            lines = [l for l in text.splitlines() if l.strip()]
            rough = [[l] for l in lines]
            cleaned.append(rough)
        all_tables.extend(cleaned)
    return all_tables

# ---------------------- Header → years map ---------------------

def map_header_to_years(header_cells: List[str], reference_year: int) -> List[Optional[int]]:
    # 1) Explicitní roky
    out: List[Optional[int]] = []
    found_any_year = False
    for c in header_cells:
        m = YEAR_RX.search(nz(c))
        if m:
            out.append(int(m.group(1)))
            found_any_year = True
        else:
            out.append(None)
    if found_any_year:
        return out
    # 2) Role "běžné/minulé"
    roles = []
    for c in header_cells:
        txt = nz(c)
        if LABEL_CURRENT.search(txt):
            roles.append("curr")
        elif LABEL_PRIOR.search(txt):
            roles.append("prev")
        else:
            roles.append(None)
    mapped: List[Optional[int]] = []
    prev_seen = 0
    for r in roles:
        if r == "curr":
            mapped.append(reference_year)
        elif r == "prev":
            mapped.append(reference_year - 1 - prev_seen)
            prev_seen += 1
        else:
            mapped.append(None)
    return mapped

# ---------------------- Harvest values -------------------------

def best_match(label: str, choices: List[str]) -> Tuple[str, int]:
    res = process.extractOne(label, choices, scorer=fuzz.WRatio)
    if res is None:
        return ("", 0)
    return (res[0], int(res[1]))

def harvest_items(tables: List[List[List[str]]], ref_year: int) -> Dict[int, Dict[str, float]]:
    """Vrací {year: {normalized_label: value}}"""
    by_year: Dict[int, Dict[str, float]] = {}

    def setv(y: int, lbl: str, val: float):
        if y not in by_year:
            by_year[y] = {}
        by_year[y][norm(lbl)] = val

    for t in tables:
        if not t:
            continue
        header = t[0]
        header_years = map_header_to_years(header, ref_year)
        for row in t[1:]:
            if not row or len(row) < 2:
                continue
            raw_label = row[0]
            label = apply_synonyms(clean_label(raw_label))
            if not LETTERS_RX.search(label) or len(label) < 3:
                continue
            nums = [to_number(c) for c in row[1:]]
            if all(v is None for v in nums):
                continue
            for idx, val in enumerate(nums):
                if val is None:
                    continue
                y = header_years[idx] if idx < len(header_years) else None
                if y is None:
                    # heuristika: první numerický sloupec = R, další = R-1, atd.
                    y = ref_year - idx
                setv(int(y), label, float(val))
    return by_year

# --------------------------- API -------------------------------

@app.post("/process")
async def process_pdf(
    file: UploadFile = File(..., description="PDF VZZ; název obsahuje rok, např. 2023.pdf"),
    template: UploadFile = File(..., description="XLSX šablona (List1), D = názvy položek, E–J = roky"),
    threshold: int = Form(76),
    return_log: bool = Form(True),
):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Soubor 'file' musí být PDF.")
    if not template.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Soubor 'template' musí být XLSX.")

    # 1) Referenční rok z názvu souboru
    m = YEAR_RX.search(file.filename)
    if not m:
        raise HTTPException(400, "V názvu PDF jsem nenašel rok (např. 2023.pdf).")
    ref_year = int(m.group(1))

    pdf_bytes = await file.read()
    tpl_bytes = await template.read()

    # 2) Načti šablonu a známé roky/sloupce
    wb, ws, year_cols, known_years, template_labels = collect_template_structure(tpl_bytes)
    if not year_cols:
        raise HTTPException(400, "V šabloně jsem nenašel roky v E–J na řádku 3.")

    log_rows: List[List[str]] = []
    def log(msg: str):
        log_rows.append([msg])

    if ref_year not in year_cols:
        log(f"Upozornění: rok {ref_year} není v hlavičce šablony (E–J). Položky pro tento rok přeskočím.")
    if (ref_year - 1) not in year_cols:
        log(f"Poznámka: rok {ref_year-1} (minulé období) není v šabloně – pokud bude v PDF, přeskočím.")

    # 3) PDF → tabulky
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages = detect_vzz_pages(pdf)
        tables = extract_tables_from_pages(pdf, pages)

    # 4) Vytěž hodnoty a roky podle hlaviček vs. ref_year
    by_year = harvest_items(tables, ref_year)

    # 5) Fuzzy párování názvů na D sloupec a zápis
    template_keys = list(template_labels.keys())
    filled_any = False
    skipped = 0
    for year, items in by_year.items():
        target_col = year_cols.get(year)
        if target_col is None:
            for lbl, val in items.items():
                skipped += 1
                log(f"Rok {year} není v šabloně – přeskočeno: '{lbl}' = {val}")
            continue
        for label_pdf, value in items.items():
            mlbl, score = best_match(norm(label_pdf), template_keys)
            if score >= threshold:
                row_idx = template_labels[mlbl]
                ws[f"{target_col}{row_idx}"].value = value
                filled_any = True
            else:
                skipped += 1
                log(f"Neshoda (<{threshold}): '{label_pdf}' ~ '{mlbl}' ({score}) → {value}")

    if not filled_any:
        log("Varování: Nepodařilo se zapsat žádnou hodnotu nad zadaný threshold.")

    # 6) LOG sheet
    if return_log:
        if "_LOG" in [s.title for s in wb.worksheets]:
            ws_log = wb["_LOG"]
        else:
            ws_log = wb.create_sheet("_LOG")
        ws_log.append([f"Zdroj: {file.filename}"])
        ws_log.append([f"Referenční rok (běžné období): {ref_year}"])
        ws_log.append([f"Prahová shoda: {threshold}"])
        ws_log.append([""])
        ws_log.append(["Poznámky / nespárované položky:"])
        for r in log_rows:
            ws_log.append(r)

    # 7) Odevzdej vyplněnou šablonu
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vyplnena_sablona_{ref_year}.xlsx"}
    )

# ---------------------- Requirements (for reference) ----------------------
# fastapi
# uvicorn[standard]
# pdfplumber
# openpyxl
# rapidfuzz
# python-multipart

